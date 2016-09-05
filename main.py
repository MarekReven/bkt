#-*- coding: utf-8 -*-
from PyQt5 import QtGui
from PyQt5 import QtWidgets# Import the PyQt4 module we'll need
import sys  # We need sys so that we can pass argv to QApplication
import design  # This file holds our MainWindow and all design related things
import csv, sqlite3
from openpyxl import load_workbook




class ExampleApp(QtWidgets.QMainWindow, design.Ui_MainWindow):


    def __init__(self):

        super(self.__class__, self).__init__()
        self.setupUi(self)

        #list of comboboBoxes in layout
        #need to be exaclty the same as in GUI
        #need to be exactly the same as layout of tables in data_sheet.csv
        # comboboxes_list = [self.comboBox, self.comboBox_2, self.comboBox_3, self.comboBox_4, self.comboBox_5,
        #                    self.comboBox_6, self.comboBox_7, self.comboBox_8, self.comboBox_9]
        comboboxes_list = [self.comboBox, self.comboBox_10, self.comboBox_11, self.comboBox_2,
                           self.comboBox_3,
                           self.comboBox_12, self.comboBox_4, self.comboBox_5,
                           self.comboBox_6, self.comboBox_7, self.comboBox_8, self.comboBox_9]


        labels_list = [self.model, self.akcesoria, self.typy_czujnikow, self.uchwyty, self.zasilanie, self.dlugosc,
                       self.kontrola,
                       self.ochrona, self.ilosc_1, self.typ_1, self.ilosc_2,
                       self.typ_2, self.wymiary, self.dobrany_kabel, self.index_short, self.index_long]
        dict_list = self.populate_data_in_dicts('data_sheet.csv')
        self.fill_comboBoxes(dict_list, comboboxes_list)
        self.make_widgets_signals(dict_list, comboboxes_list)
        self.pushButton.clicked.connect(lambda: self.update_xlsx('choosen_data.xlsx',
                                                                 comboboxes_list, labels_list))
        #set defaul value to first combo (model)
        self.comboBox.setCurrentIndex(1)

        #set defaul value for cabel combo
        self.comboBox_12.setCurrentIndex(1)



    def btnstate(self):
        # print(1)
        for radioButton in self.findChildren(QtWidgets.QRadioButton):
            # print(2)
            if radioButton.isChecked():
                print(radioButton.parent())
                radioButtonText = radioButton.text()
                print(radioButtonText)


    def calculate_max_load(self):
        # get txt ffrom comboBox 'wtyk zasilajacy'
        plug_val = self.comboBox_3.currentText()

        voltage = plug_val.split('A', 1)[1][1:4]
        amperage = plug_val.split('A', 1)[0][-2:]

        if voltage == '250':
            max_load = '{0}A'.format(amperage)
        elif voltage == '400':
            max_load = '3x{0}A'.format(amperage)
        else:
            print('blad')

        return max_load

    def calculate_max_power(self):

        plug = self.comboBox_3.currentText()
        amperage = plug.split('A', 1)[0][-2:]
        voltage = plug.split('A', 1)[1][1:4]
        power = str(int(amperage) * 230)

        if voltage == '250':
            max_power = '{0}W'.format(power)
        elif voltage == '400':
            max_power = '3x{0}W'.format(power)
        else:
            print('blad')


        return max_power

    def sum_additional_elements(self):
        all_elements = ''
        all_elements = '{0}, {1}'.format(self.comboBox_4.currentText(), self.comboBox_5.currentText())
        model_name = self.comboBox.currentText()
        if model_name == 'NPM-V':
            all_elements += ', {0}, {1}'.format(self.comboBox_10.currentText(), self.comboBox_11.currentText())
        elif model_name == 'IP-PDU':
            all_elements += ', dodatkowe gniazdo do podłączenia czujnika 1 x temp/wilgotności'
        else:
            pass
        return all_elements


    def update_xlsx(self, dest, comboboxes_list, labels_list):
        # Open an xlsx for reading
        wb = load_workbook(filename=dest)
        # Get the current Active Sheet
        ws = wb.get_active_sheet()
        count_a = 0

        #get labels
        for label in labels_list:
            count_a += 1
            ws.cell('A{0}'.format(count_a)).value = label.text()

        #get comboboxesvalues
        count_b = 0
        for combo in comboboxes_list:
            count_b += 1
            ws.cell('B{0}'.format(count_b)).value = combo.currentText()
        #get linesvalues
        count_b += 1
        ws.cell('B{0}'.format(count_b)).value = self.lineEdit_5.text()
        count_b += 1
        ws.cell('B{0}'.format(count_b)).value = self.lineEdit_4.text()
        count_b += 1
        ws.cell('B{0}'.format(count_b)).value = self.lineEdit.text()
        count_b += 1
        ws.cell('B{0}'.format(count_b)).value = self.lineEdit_2.text()

        #max obciazenie
        count_b += 1
        ws.cell('A{0}'.format(count_b)).value = "Maksymalne obciążenie:"
        ws.cell('B{0}'.format(count_b)).value = self.calculate_max_load()
        #moc znamionowa
        count_b += 1
        ws.cell('A{0}'.format(count_b)).value = "Moc znamionowa:"
        ws.cell('B{0}'.format(count_b)).value = self.calculate_max_power()
        #elementy dodatkowe razem
        count_b += 1
        ws.cell('A{0}'.format(count_b)).value = "Elementy dodatkowe łącznie:"
        ws.cell('B{0}'.format(count_b)).value = self.sum_additional_elements()

        #if cable length not in standard
        if self.comboBox_12.currentText() == "Niestandardowy":
            ws.cell('B6').value = self.lineEdit_3.text()
        else:
            pass

        wb.save('choosen_data.xlsx')

    def populate_data_in_dicts(self, csv_file):
        #DICTIONARIES
        model_listwy = {}
        akcesoria_npmv = {}
        typy_czujnikow = {}
        uchwyty = {}
        zasilanie_na_wejsciu = {}
        dlugosc = {}
        kontrola_pdu = {}
        ochrona_pdu = {}
        ilosc_gniazd_i_typu = {}
        i_typ_gniazd = {}
        ilosc_gniazd_ii_typu = {}
        ii_typ_gniazd = {}

        dicts_list = [model_listwy,
                      akcesoria_npmv,
                      typy_czujnikow,
                      uchwyty,
                      zasilanie_na_wejsciu,
                      dlugosc, kontrola_pdu,
                      ochrona_pdu,
                      ilosc_gniazd_i_typu,
                      i_typ_gniazd,
                      ilosc_gniazd_ii_typu,
                      ii_typ_gniazd]

        csv_lines = []
        with open(csv_file, 'r') as f:
            reader = csv.reader(f)
            for row in reader:
                csv_lines.append(row)
            # print(csv_lines)

            for dictionary in dicts_list:
                count = 0
                for item in csv_lines:
                    count += 1
                    if count > 2:
                        dictionary[item[0]] = item[1]
                        del item[0:3]
                    else:
                        pass
        return dicts_list

    def fill_comboBoxes(self, dicts_list, comboboxes_list):
        i = 0
        for combo in comboboxes_list:
            combo.addItems(sorted(dicts_list[i].keys()))
            i += 1

    def make_widgets_signals(self,dicts_list, comboboxes_list):

        for combo in comboboxes_list:
            combo.currentIndexChanged.connect(lambda: self.make_indexes(dicts_list, comboboxes_list))

        for radioButton in self.findChildren(QtWidgets.QRadioButton):
            radioButton.toggled.connect(lambda: self.make_indexes(dicts_list, comboboxes_list))

    def enable_disable(self, widget_to_disable, combo_to_check, enable_on_this_string):

        if combo_to_check.currentText() == enable_on_this_string:
            widget_to_disable.setEnabled(True)
        else:
            widget_to_disable.setEnabled(False)

    def show_hide(self, widget_to_hide, combo_to_check, show_on_this_string):

        if combo_to_check.currentText() == show_on_this_string:
            widget_to_hide.show()
        else:
            widget_to_hide.hide()

    def make_indexes(self, dicts_list, comboboxes_list):

        #show/hide enable/disable widgets on signal from comboboxes
        self.enable_disable(self.groupBox, self.comboBox, "NPM-V")
        self.show_hide(self.frame, self.comboBox_12, "Niestandardowy")


        cable_index = ''
        long_index = ''
        short_index = ''

        #----Cable index calculation

        #cable_type
        plug = self.comboBox_3.currentText()
        if '32A/400V' in plug:
            cable_type = '5x6.0mm2'
        elif '32A/250V' in plug:
            cable_type = '3x6.0mm2'
        elif '16A/250V' in plug:
            if '60309' in plug:
                cable_type = '3x2.5mm2'
            elif '320' in plug:
                cable_type = '3x1.5mm2'
            else:
                cable_type = 'blad'
        elif '10A/250V' in plug:
            cable_type = '3x1.5mm2'
        elif '63A/400V' in plug:
            cable_type = '5x10.0mm2'
        elif '16A/400V' in plug:
            cable_type = '5x2.5mm2'
        else:
            cable_type = '3x1.5mm2'

        #cable_length
        cable_length = ''
        cable_combo_text = self.comboBox_12.currentText()
        if cable_combo_text == 'Niestandardowy':
            def te_xt_changed(cable_type):
                print('here')
                cable_length = (self.lineEdit_3.text())
                cable_index = ''
                cable_index += 'H05W-F {0}, {1}, czarny'.format(cable_type, cable_length)
                self.lineEdit_4.setText(cable_index)
            self.lineEdit_3.textChanged.connect(lambda: te_xt_changed(cable_type))
        elif cable_combo_text == 'Brak':
            cable_index = 'Brak'
        else:
            cable_length = self.comboBox_12.currentText()

        if cable_index != 'Brak':
            cable_index += 'H05W-F {0}, {1}, czarny'.format(cable_type, cable_length)
        else:
            cable_index = 'listwa bez kabla'

        #set cable value
        self.lineEdit_4.setText(cable_index)


        #Short Index
        #get values from radiobuttons
        radio_button_values = []
        count_radio = 0
        for radioButton in self.findChildren(QtWidgets.QRadioButton):
            # print(2)
            if radioButton.isChecked():
                count_radio += 1
                print(radioButton.parent())
                radioButtonText = radioButton.text()
                print(radioButtonText)
                radio_button_values.append(radioButtonText)
                if count_radio == 2:
                    radioValue1 = radio_button_values[0]
                    radioValue2 = radio_button_values[1]

        # #make short index
        count = 0
        short_index = ''
        for combo, dict in zip(comboboxes_list, dicts_list):
            # short_index = ''
            count += 1
            if self.comboBox.currentText() == "NPM-V":
                # short_index = ''
                # count += 1
                if count == 2:
                    akcesoria_dodatkowe = dict[combo.currentText()]
                else:
                    short_index += dict[combo.currentText()]
                    if count == 5 and count_radio >= 2:
                        short_index += '{0}.{1}{2}'.format(radioValue1, radioValue2, akcesoria_dodatkowe)
                        print(dict[combo.currentText()])
                    elif count == 9:
                        short_index += '.'
                    elif count == 10:
                        short_index += '-'
                    elif count == 11:
                        short_index += ','
                    elif count == 12:
                        short_index += '-'
            else:
                #clear combo of NPM-V accessories to not influence short_index
                self.comboBox_10.setCurrentIndex(-1)
                short_index += dict[combo.currentText()]
                if count == 8:
                    short_index += '.'
                    print(dict[combo.currentText()])
                elif count == 9:
                    short_index += '-'
                elif count == 10:
                    short_index += ','
                elif count == 11:
                    short_index += '-'
            self.lineEdit.setText(short_index)

        #Long index
        if self.comboBox.currentText() == 'PDU - Basic':
            zero = 'zasilająca'
        else:
            zero = 'zarządzalna'
        # print(self.comboBox_5.currentText())
        long_index += 'Listwa {0} pionowa'.format(zero)
        if '400' in self.comboBox_3.currentText():
            long_index += ' {0}'.format('trójfazowa')
        else:
            pass
        long_index += ' BKT {0} typ {1} x {2}'.format(self.comboBox.currentText(),
                                                     self.comboBox_6.currentText(),
                                                     self.comboBox_7.currentText())
        # print(type(self.comboBox_8.currentText()))
        if self.comboBox_8.currentText() != '0':
            long_index += ' + {0} x {1}'.format(self.comboBox_8.currentText(), self.comboBox_9.currentText())
        else:
            pass
        long_index += ' , {0}'.format(self.comboBox_3.currentText())

        self.lineEdit_2.setText(long_index)


def main():
    app = QtWidgets.QApplication(sys.argv)
    form = ExampleApp()
    form.show()
    app.exec_()  # and execute the app


if __name__ == '__main__':
    main()